from __future__ import annotations

import asyncio
import base64
import hmac
import json
import logging
import os
import secrets
import uuid
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from dataclasses import dataclass
from typing import Any, AsyncGenerator, Dict, List, Optional, Sequence, Set, Tuple, Type, TypeVar

from hashlib import pbkdf2_hmac
from jose import ExpiredSignatureError, JWTError, jwt as PyJWT
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, FastAPI, HTTPException, status
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer

import httpx

from pydantic import BaseModel, ConfigDict, EmailStr, Field, field_validator
from sqlalchemy import (
    JSON,
    Boolean,
    DateTime,
    ForeignKey,
    Integer,
    String,
    Text,
    UniqueConstraint,
    delete,
    func,
    inspect,
    select,
    text,
)
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column
from starlette.middleware.cors import CORSMiddleware

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

app = FastAPI()

_allowed_origins = os.environ.get("CORS_ALLOW_ORIGINS", "*")
allow_origin_list = [origin.strip() for origin in _allowed_origins.split(",") if origin.strip()]
if "*" in allow_origin_list:
    allow_origin_list = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origin_list,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
api_router = APIRouter(prefix="/api")
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# ==================== DATABASE SETUP ====================

DATABASE_URL = os.environ.get(
    "DATABASE_URL", f"sqlite+aiosqlite:///{(ROOT_DIR / 'app.db').as_posix()}"
)

engine = create_async_engine(DATABASE_URL, echo=False, future=True)
async_session = async_sessionmaker(engine, expire_on_commit=False)

PASSWORD_HASH_ALGORITHM = "pbkdf2_sha256"
PASSWORD_HASH_ITERATIONS = int(os.environ.get("PASSWORD_HASH_ITERATIONS", "390000"))


def _b64encode(data: bytes) -> str:
    return base64.b64encode(data).decode("utf-8")


def _b64decode(data: str) -> bytes:
    return base64.b64decode(data.encode("utf-8"))


def hash_password(password: str) -> str:
    if not isinstance(password, str):
        raise ValueError("Password must be a string")
    salt = secrets.token_bytes(16)
    derived = pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        PASSWORD_HASH_ITERATIONS,
    )
    return "$".join(
        (
            PASSWORD_HASH_ALGORITHM,
            str(PASSWORD_HASH_ITERATIONS),
            _b64encode(salt),
            _b64encode(derived),
        )
    )


def is_password_hash(value: str) -> bool:
    if not isinstance(value, str):
        return False
    parts = value.split("$")
    if len(parts) != 4 or parts[0] != PASSWORD_HASH_ALGORITHM:
        return False
    try:
        int(parts[1])
        _b64decode(parts[2])
        _b64decode(parts[3])
    except Exception:
        return False
    return True


def verify_password(password: str, stored_hash: str) -> bool:
    if not stored_hash:
        return False
    if is_password_hash(stored_hash):
        algorithm, iterations_str, salt_b64, hash_b64 = stored_hash.split("$")
        if algorithm != PASSWORD_HASH_ALGORITHM:
            return False
        try:
            iterations = int(iterations_str)
        except ValueError:
            return False
        salt = _b64decode(salt_b64)
        expected = _b64decode(hash_b64)
        candidate = pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            salt,
            iterations,
        )
        return hmac.compare_digest(expected, candidate)
    return password == stored_hash


class Base(DeclarativeBase):
    pass


class TimestampMixin:
    created_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True), default=lambda: datetime.now(timezone.utc)
    )


class UserTable(Base, TimestampMixin):
    __tablename__ = "users"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    email: Mapped[str] = mapped_column(String, unique=True, index=True, nullable=False)
    username: Mapped[str] = mapped_column(String, nullable=False)
    role: Mapped[str] = mapped_column(String, nullable=False)
    password_hash: Mapped[str] = mapped_column(String, nullable=False)


class PlanCycleTable(Base, TimestampMixin):
    __tablename__ = "plan_cycles"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    name: Mapped[str] = mapped_column(String, unique=True, nullable=False)
    description: Mapped[Optional[str]] = mapped_column(Text, nullable=True)


class ProjectTable(Base, TimestampMixin):
    __tablename__ = "projects"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    plan_cycle_id: Mapped[str] = mapped_column(
        String,
        ForeignKey("plan_cycles.id", ondelete="CASCADE"),
        index=True,
        nullable=False,
    )
    name: Mapped[str] = mapped_column(String, nullable=False)
    description: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    created_by: Mapped[str] = mapped_column(String, nullable=False)


class ProjectLinkedMixin:
    plan_cycle_id: Mapped[str] = mapped_column(
        String,
        ForeignKey("plan_cycles.id", ondelete="CASCADE"),
        index=True,
        nullable=False,
    )
    project_id: Mapped[str] = mapped_column(String, index=True, nullable=False)


class ProjectAccessTable(Base, TimestampMixin):
    __tablename__ = "project_access"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    user_id: Mapped[str] = mapped_column(String, index=True, nullable=False)
    plan_cycle_id: Mapped[str] = mapped_column(
        String,
        ForeignKey("plan_cycles.id", ondelete="CASCADE"),
        index=True,
        nullable=False,
    )
    project_id: Mapped[str] = mapped_column(String, index=True, nullable=False)
    visible: Mapped[bool] = mapped_column(Boolean, nullable=False, default=True)

    __table_args__ = (
        UniqueConstraint(
            "user_id",
            "plan_cycle_id",
            "project_id",
            name="uq_project_access",
        ),
    )


class RevisionHistoryTable(Base, ProjectLinkedMixin):
    __tablename__ = "revision_history"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    revision_no: Mapped[str] = mapped_column(String, nullable=False)
    date: Mapped[str] = mapped_column(String, nullable=False)
    change_description: Mapped[str] = mapped_column(Text, nullable=False)
    authors: Mapped[str] = mapped_column(Text, nullable=False)
    reviewed_by: Mapped[str] = mapped_column(String, nullable=False)
    approved_by: Mapped[str] = mapped_column(String, nullable=False)
    

class TOCEntryTable(Base, ProjectLinkedMixin):
    __tablename__ = "toc_entries"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    sheet_name: Mapped[str] = mapped_column(String, nullable=False)
    sections_in_sheet: Mapped[str] = mapped_column(Text, nullable=False)


class DefinitionAcronymTable(Base, ProjectLinkedMixin):
    __tablename__ = "definition_acronyms"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    term: Mapped[str] = mapped_column(String, nullable=False)
    definition: Mapped[str] = mapped_column(Text, nullable=False)


class SingleEntryFieldTable(Base, ProjectLinkedMixin):
    __tablename__ = "single_entry_fields"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    field_name: Mapped[str] = mapped_column(String, nullable=False)
    content: Mapped[str] = mapped_column(Text, nullable=False)
    image_data: Mapped[Optional[str]] = mapped_column(Text, nullable=True)


class ProjectDetailsTable(Base, ProjectLinkedMixin):
    __tablename__ = "project_details"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    project_model: Mapped[str] = mapped_column(String, nullable=False)
    project_type: Mapped[str] = mapped_column(String, nullable=False)
    software_type: Mapped[str] = mapped_column(String, nullable=False)
    standard_to_be_followed: Mapped[str] = mapped_column(String, nullable=False)
    customer: Mapped[str] = mapped_column(String, nullable=False)
    programming_language: Mapped[str] = mapped_column(String, nullable=False)
    project_duration: Mapped[str] = mapped_column(String, nullable=False)
    team_size: Mapped[str] = mapped_column(String, nullable=False)


class AssumptionTable(Base, ProjectLinkedMixin):
    __tablename__ = "assumptions"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    sl_no: Mapped[str] = mapped_column(String, nullable=False)
    brief_description: Mapped[str] = mapped_column(Text, nullable=False)
    impact_on_project_objectives: Mapped[str] = mapped_column(Text, nullable=False)
    remarks: Mapped[Optional[str]] = mapped_column(Text, nullable=True)


class ConstraintTable(Base, ProjectLinkedMixin):
    __tablename__ = "constraints"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    constraint_no: Mapped[str] = mapped_column(String, nullable=False)
    brief_description: Mapped[str] = mapped_column(Text, nullable=False)
    impact_on_project_objectives: Mapped[str] = mapped_column(Text, nullable=False)
    remarks: Mapped[Optional[str]] = mapped_column(Text, nullable=True)


class DependencyTable(Base, ProjectLinkedMixin):
    __tablename__ = "dependencies"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    sl_no: Mapped[str] = mapped_column(String, nullable=False)
    brief_description: Mapped[str] = mapped_column(Text, nullable=False)
    impact_on_project_objectives: Mapped[str] = mapped_column(Text, nullable=False)
    remarks: Mapped[Optional[str]] = mapped_column(Text, nullable=True)


class StakeholderTable(Base, ProjectLinkedMixin):
    __tablename__ = "stakeholders"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    sl_no: Mapped[str] = mapped_column(String, nullable=False)
    name: Mapped[str] = mapped_column(String, nullable=False)
    stakeholder_type: Mapped[str] = mapped_column(String, nullable=False)
    role: Mapped[str] = mapped_column(String, nullable=False)
    authority_responsibility: Mapped[str] = mapped_column(Text, nullable=False)
    contact_details: Mapped[str] = mapped_column(Text, nullable=False)


class DeliverableTable(Base, ProjectLinkedMixin):
    __tablename__ = "deliverables"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    sl_no: Mapped[str] = mapped_column(String, nullable=False)
    work_product: Mapped[str] = mapped_column(String, nullable=False)
    owner_of_deliverable: Mapped[str] = mapped_column(String, nullable=False)
    approving_authority: Mapped[str] = mapped_column(String, nullable=False)
    release_to_customer: Mapped[str] = mapped_column(String, nullable=False)
    milestones: Mapped[Dict[str, str]] = mapped_column(JSON, default=dict)


class MilestoneColumnTable(Base, ProjectLinkedMixin):
    __tablename__ = "milestone_columns"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    column_name: Mapped[str] = mapped_column(String, nullable=False)
    order: Mapped[int] = mapped_column(Integer, nullable=False)


class SamDeliverableTable(Base, ProjectLinkedMixin):
    __tablename__ = "sam_deliverables"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    sl_no: Mapped[str] = mapped_column(String, nullable=False)
    work_product: Mapped[str] = mapped_column(String, nullable=False)
    owner_of_deliverable: Mapped[str] = mapped_column(String, nullable=False)
    approving_authority: Mapped[str] = mapped_column(String, nullable=False)
    release_to_tsbj: Mapped[str] = mapped_column(String, nullable=False)
    milestones: Mapped[Dict[str, str]] = mapped_column(JSON, default=dict)


class SamMilestoneColumnTable(Base, ProjectLinkedMixin):
    __tablename__ = "sam_milestone_columns"

    id: Mapped[str] = mapped_column(
        String, primary_key=True, default=lambda: str(uuid.uuid4())
    )
    column_name: Mapped[str] = mapped_column(String, nullable=False)
    order: Mapped[int] = mapped_column(Integer, nullable=False)


@dataclass(frozen=True)
class ColumnDefinition:
    name: str
    type_: Any = Text
    nullable: bool = True


@dataclass(frozen=True)
class SectionTableDefinition:
    section: str
    key: str
    columns: Sequence[ColumnDefinition]
    table_name: Optional[str] = None


def _to_camel_case(value: str) -> str:
    return "".join(part.capitalize() for part in value.split("_"))


def _create_section_table_model(definition: SectionTableDefinition) -> Type[ProjectLinkedMixin]:
    table_name = definition.table_name or definition.key
    class_name = f"{definition.section}{_to_camel_case(definition.key)}Table"

    annotations: Dict[str, Any] = {"id": Mapped[str]}
    attrs: Dict[str, Any] = {
        "__tablename__": table_name,
        "__annotations__": annotations,
        "id": mapped_column(
            String, primary_key=True, default=lambda: str(uuid.uuid4())
        ),
    }

    for column in definition.columns:
        python_type = Optional[str] if column.nullable else str
        annotations[column.name] = Mapped[python_type]
        attrs[column.name] = mapped_column(column.type_, nullable=column.nullable)

    model_cls = type(class_name, (Base, ProjectLinkedMixin), attrs)
    globals()[class_name] = model_cls
    return model_cls


def col(name: str, type_: Any = Text, nullable: bool = True) -> ColumnDefinition:
    return ColumnDefinition(name=name, type_=type_, nullable=nullable)


@dataclass(frozen=True)
class SectionTableMeta:
    model: Type[ProjectLinkedMixin]
    columns: Sequence[str]


SECTION_TABLE_DEFINITIONS: Sequence[SectionTableDefinition] = [
    SectionTableDefinition(
        section="M4",
        key="business_continuity",
        columns=[
            col("sl_no", String, False),
            col("brief_description"),
            col("impact_of_project_objectives"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M4",
        key="information_security_requirements",
        columns=[
            col("sl_no", String, False),
            col("phase", String),
            col("is_requirement_description"),
            col("monitoring_control"),
            col("tools"),
            col("artifacts"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M5",
        key="human_resource_and_special_training_plan",
        columns=[
            col("sl_no", String, False),
            col("role", String),
            col("skill_experience_required"),
            col("no_of_people_required", String),
            col("available", String),
            col("project_specific_training_needs"),
        ],
    ),
    SectionTableDefinition(
        section="M5",
        key="environment_and_tools",
        columns=[
            col("sl_no", String, False),
            col("name_brief_description"),
            col("no_of_licenses_required", String),
            col("source", String),
            col("status", String),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M5",
        key="build_buy_reuse",
        columns=[
            col("sl_no", String, False),
            col("component_product"),
            col("build_buy_reuse", String),
            col("reuse_goals_objectives"),
            col("vendor_project_name_version", String),
            col("responsible_person_reuse", String),
            col("quality_evaluation_criteria"),
            col("responsible_person_qualification"),
            col("modifications_planned"),
            col("selected_item_operational_environment"),
            col("known_defect_vulnerabilities_limitations"),
        ],
    ),
    SectionTableDefinition(
        section="M5",
        key="reuse_analysis",
        columns=[
            col("sl_no", String, False),
            col("component_product"),
            col("reuse", String),
            col("modifications_required"),
            col("constraints_for_reuse"),
            col("risk_analysis_result"),
            col("impact_on_plan_activities"),
            col("evaluation_to_comply_cyber_security"),
            col("impact_on_integration_documents"),
            col("known_defects"),
        ],
    ),
    SectionTableDefinition(
        section="M5",
        key="size_and_complexity",
        columns=[
            col("sl_no", String, False),
            col("product_component_module"),
            col("size_kloc", String),
            col("percent_reuse_estimated", String),
            col("effort_person_days_weeks_months", String),
            col("complexity", String),
        ],
    ),
    SectionTableDefinition(
        section="M5",
        key="duration_effort_estimate_organization_norms",
        columns=[
            col("sl_no", String, False),
            col("phase_milestone", String),
            col("schedule_days_weeks", String),
            col("effort_in_per_days_weeks", String),
            col("remarks_on_deviation"),
        ],
    ),
    SectionTableDefinition(
        section="M5",
        key="usage_of_off_the_shelf_component",
        columns=[
            col("sl_no", String, False),
            col("name_of_component"),
            col("requirements_complied", String),
            col("requirement_document_updated", String),
            col("specific_application_context"),
            col("documentation_sufficient", String),
            col("vulnerabilities_identified"),
            col("integration_document_updated", String),
            col("test_design_document"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M5",
        key="cybersecurity_interface_agreement",
        columns=[
            col("sl_no", String, False),
            col("phase", String),
            col("work_product"),
            col("document_ref", String),
            col("supplier", String),
            col("customer", String),
            col("level_of_confidentiality", String),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M6",
        key="project_monitoring_and_control",
        columns=[
            col("sl_no", String, False),
            col("type_of_progress_reviews"),
            col("month_phase_milestone_frequency"),
            col("participants"),
            col("remarks"),
            col("mode_of_communication", String),
        ],
    ),
    SectionTableDefinition(
        section="M6",
        key="quantitative_objectives_measurement_and_data_management_plan",
        columns=[
            col("sl_no", String, False),
            col("objective"),
            col("metric"),
            col("priority", String),
            col("project_goal"),
            col("organisation_norm"),
            col("data_source"),
            col("reason_for_deviation_from_organization_norm"),
        ],
    ),
    SectionTableDefinition(
        section="M7",
        key="standards_qm",
        columns=[
            col("sl_no", String, False),
            col("name_of_standard"),
            col("brief_description"),
            col("source", String),
        ],
    ),
    SectionTableDefinition(
        section="M7",
        key="verification_and_validation_plan",
        columns=[
            col("sl_no", String, False),
            col("artifact_name"),
            col("verification_method"),
            col("verification_type", String),
            col("validation_method"),
            col("validation_type", String),
            col("tools_used"),
            col("approving_authority", String),
            col("verification_validation_evidence"),
            col("remarks_deviation"),
        ],
    ),
    SectionTableDefinition(
        section="M7",
        key="confirmation_review_plan",
        columns=[
            col("sl_no", String, False),
            col("artifact_name"),
            col("phase", String),
            col("confirmation_measure"),
            col("plan_schedule"),
            col("asil", String),
            col("independence_level", String),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M7",
        key="proactive_causal_analysis_plan",
        columns=[
            col("sl_no", String, False),
            col("previous_similar_projects_executed"),
            col("major_issues_defects_identified_by_customer"),
            col("corrective_preventive_measures"),
        ],
    ),
    SectionTableDefinition(
        section="M7",
        key="reactive_causal_analysis_plan",
        columns=[
            col("sl_no", String, False),
            col("phase_milestone", String),
            col("brief_description_of_instances_when_causal_analysis_needs_to_be_done"),
            col("causal_analysis_method_tool"),
            col("responsibility", String),
        ],
    ),
    SectionTableDefinition(
        section="M8",
        key="decision_management_plan",
        columns=[
            col("sl_no", String, False),
            col("phase_milestone", String),
            col("brief_description_of_major_decisions"),
            col("decision_making_method_tool"),
            col("responsibility", String),
        ],
    ),
    SectionTableDefinition(
        section="M8",
        key="tailoring_qms",
        columns=[
            col("sl_no", String, False),
            col("brief_description_of_deviation"),
            col("gap_analysis_details"),
            col("reasons_justifications"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M8",
        key="deviations",
        columns=[
            col("sl_no", String, False),
            col("brief_description_of_deviation"),
            col("reasons_justifications"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M8",
        key="product_release_plan",
        columns=[
            col("sl_no", String, False),
            col("release_type", String),
            col("objective"),
            col("release_date_milestones"),
            col("mode_of_delivery", String),
            col("qa_release_audit_date", String),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M8",
        key="tailoring_due_to_component_out_of_context",
        columns=[
            col("sl_no", String, False),
            col("phase_milestone", String),
            col("brief_description_of_major_decisions"),
            col("decision_making_method_tool"),
            col("responsibility", String),
        ],
    ),
    SectionTableDefinition(
        section="M8",
        key="release_cybersecurity_interface_agreement",
        columns=[
            col("sl_no", String, False),
            col("phase", String),
            col("work_product"),
            col("document_ref", String),
            col("supplier", String),
            col("customer", String),
            col("level_of_confidentiality", String),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M9",
        key="risk_management_plan",
        columns=[
            col("sl_no", String, False),
            col("risk_identification_method"),
            col("phase_sprint_milestone"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M9",
        key="risk_mitigation_and_contingency",
        columns=[
            col("risk_id", String, False),
            col("risk_description"),
            col("risk_category", String),
            col("risk_originator_name", String),
            col("risk_source", String),
            col("date_of_risk_identification", String),
            col("phase_of_risk_identification", String),
            col("risk_treatment_option", String),
            col("rationale_to_choose_risk_treatment_option"),
            col("effort_required_for_risk_treatment"),
            col("risk_treatment_schedule"),
            col("success_criteria_for_risk_treatment_activities"),
            col("criteria_for_cancellation_of_risk_treatment_activities"),
            col("frequency_of_monitoring_risk_treatment_activities"),
            col("threshold"),
            col("trigger"),
            col("probability", String),
            col("impact", String),
            col("risk_exposure", String),
            col("mitigation_plan"),
            col("contingency_plan"),
            col("verification_methods_for_mitigation_contingency_plan"),
            col("list_of_stakeholders"),
            col("responsibility", String),
            col("status", String),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M9",
        key="risk_exposure_history",
        columns=[
            col("risk", String),
            col("date", String),
            col("exposure_value", String),
        ],
    ),
    SectionTableDefinition(
        section="M10",
        key="opportunity_register",
        columns=[
            col("opportunity_id", String, False),
            col("opportunity_description"),
            col("opportunity_category", String),
            col("opportunity_source", String),
            col("date_of_identification", String),
            col("phase_of_identification", String),
            col("cost", String),
            col("benefit", String),
            col("opportunity_value", String),
            col("leverage_plan_to_maximize_opportunities_identified"),
            col("responsibility", String),
            col("status", String),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M10",
        key="opportunity_management_plan",
        columns=[
            col("sl_no", String, False),
            col("opportunity_identification_method"),
            col("phase_sprint_milestone"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M10",
        key="opportunity_value_history",
        columns=[
            col("opportunity", String),
            col("date", String),
            col("opportunity_value", String),
        ],
    ),
    SectionTableDefinition(
        section="M11",
        key="list_of_configuration_items",
        columns=[
            col("sl_no", String, False),
            col("ci_name_description"),
            col("source", String),
            col("format_type", String),
            col("description_of_level"),
            col("branching_merging_required", String),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M11",
        key="list_of_non_configurable_items",
        columns=[
            col("sl_no", String, False),
            col("ci_name_description"),
            col("source", String),
            col("format_type", String),
            col("description_of_level"),
            col("branching_merging_required", String),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M11",
        key="naming_convention",
        columns=[
            col("sl_no", String, False),
            col("files_and_folders"),
            col("naming_convention"),
            col("name_of_ci", String),
        ],
    ),
    SectionTableDefinition(
        section="M11",
        key="branching_and_merging",
        columns=[
            col("sl_no", String, False),
            col("branch_convention"),
            col("phase", String),
            col("branch_name", String),
            col("risk_associated_with_branching"),
            col("verification"),
        ],
    ),
    SectionTableDefinition(
        section="M11",
        key="labelling_baselines",
        columns=[
            col("sl_no", String, False),
            col("ci"),
            col("planned_baseline_phase_milestone_date"),
            col("criteria_for_baseline"),
            col("baseline_name_label_or_tag"),
        ],
    ),
    SectionTableDefinition(
        section="M11",
        key="labelling_baselines2",
        columns=[
            col("sl_no", String, False),
            col("branch_convention"),
            col("phase", String),
            col("branch_name_tag"),
        ],
    ),
    SectionTableDefinition(
        section="M11",
        key="configuration_control",
        columns=[
            col("sl_no", String, False),
            col("ci_or_folder_name_path"),
            col("developer_role", String),
            col("team_leader_role", String),
            col("em_role", String),
            col("ed_role", String),
            col("qa_role", String),
            col("ccb_member", String),
        ],
    ),
    SectionTableDefinition(
        section="M11",
        key="configuration_control_board",
        columns=[
            col("sl_no", String, False),
            col("ccb_members_name"),
            col("role", String),
            col("remarks_need_for_inclusion"),
        ],
    ),
    SectionTableDefinition(
        section="M11",
        key="configuration_status_accounting",
        columns=[
            col("sl_no", String, False),
            col("phase_milestone_month"),
        ],
    ),
    SectionTableDefinition(
        section="M11",
        key="configuration_management_audit",
        columns=[
            col("sl_no", String, False),
            col("phase_milestone_month"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_assumptions",
        columns=[
            col("sl_no", String, False),
            col("brief_description"),
            col("impact_on_project_objectives"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_constraints",
        columns=[
            col("constraint_no", String, False),
            col("brief_description"),
            col("impact_on_project_objectives"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_dependencies",
        columns=[
            col("sl_no", String, False),
            col("brief_description"),
            col("impact_on_project_objectives"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_risks",
        columns=[
            col("sl_no", String, False),
            col("brief_description"),
            col("impact_of_project_objectives"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_status_reporting_and_communication_plan",
        columns=[
            col("sl_no", String, False),
            col("type_of_progress_reviews"),
            col("month_phase_milestone_frequency"),
            col("participants"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_quantitative_objectives_measurement_and_data_management_plan",
        columns=[
            col("sl_no", String, False),
            col("objective"),
            col("metric", String),
            col("priority", String),
            col("project_goal"),
            col("organisation_norm"),
            col("data_source"),
            col("reason_for_deviation_from_organization_norm"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_verification_and_validation_plan",
        columns=[
            col("sl_no", String, False),
            col("work_product"),
            col("verification_method"),
            col("validation_method"),
            col("approving_authority", String),
            col("remarks_for_deviation"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="tailoring_sam",
        columns=[
            col("sl_no", String, False),
            col("brief_description_of_deviation"),
            col("reasons_justifications"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_deviations",
        columns=[
            col("sl_no", String, False),
            col("brief_description_of_deviation"),
            col("reasons_justifications"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_product_release_plan",
        columns=[
            col("sl_no", String, False),
            col("release_type", String),
            col("objective"),
            col("release_date_milestones"),
            col("remarks"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_labelling_baselines",
        columns=[
            col("sl_no", String, False),
            col("ci"),
            col("planned_baseline_phase_milestone_date"),
            col("criteria_for_baseline"),
            col("baseline_name_label_or_tag"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_labelling_baselines2",
        columns=[
            col("sl_no", String, False),
            col("branch_convention"),
            col("phase", String),
            col("branch_name_tag"),
        ],
    ),
    SectionTableDefinition(
        section="M13",
        key="sam_configuration_control",
        columns=[
            col("sl_no", String, False),
            col("ci_or_folder_name_path"),
            col("developer_role", String),
            col("team_leader_role", String),
            col("pm_role", String),
            col("pgm_dh_role", String),
            col("qa_role", String),
            col("ccb_member", String),
        ],
    ),
]


SECTION_TABLE_REGISTRY: Dict[Tuple[str, str], SectionTableMeta] = {}

for section_definition in SECTION_TABLE_DEFINITIONS:
    model_cls = _create_section_table_model(section_definition)
    SECTION_TABLE_REGISTRY[(section_definition.section, section_definition.key)] = (
        SectionTableMeta(
            model=model_cls,
            columns=[column.name for column in section_definition.columns],
        )
    )


INVALID_SHEET_TITLE_CHARS = set("[]:*?/\\")
MAX_SHEET_TITLE_LENGTH = 31
EXPORT_COLUMN_EXCLUDES = {"id", "project_id"}


def make_sheet_title(base_title: str, used_titles: Set[str]) -> str:
    sanitized = "".join(
        "-" if char in INVALID_SHEET_TITLE_CHARS else char for char in base_title
    ).strip()
    if not sanitized:
        sanitized = "Sheet"
    truncated = sanitized[:MAX_SHEET_TITLE_LENGTH]
    candidate = truncated or "Sheet"
    counter = 1
    while candidate in used_titles:
        suffix = f"_{counter}"
        candidate = f"{truncated[:MAX_SHEET_TITLE_LENGTH - len(suffix)]}{suffix}"
        counter += 1
    used_titles.add(candidate)
    return candidate


def friendly_header(column_name: str) -> str:
    return column_name.replace("_", " ").title()


def format_cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def decode_image_for_workbook(image_data: str) -> Optional[Tuple[XLImage, BytesIO]]:
    if not image_data:
        return None
    try:
        base64_data = image_data.split(",", 1)[1] if "," in image_data else image_data
        binary = base64.b64decode(base64_data)
    except Exception:
        return None

    buffer = BytesIO(binary)
    buffer.seek(0)
    try:
        image = XLImage(buffer)
    except Exception:
        return None

    max_width = 480
    max_height = 320
    width = getattr(image, "width", None)
    height = getattr(image, "height", None)
    if width and height:
        scale = min(
            1.0,
            max_width / float(width) if width else 1.0,
            max_height / float(height) if height else 1.0,
        )
        if scale < 1.0:
            image.width = int(width * scale)
            image.height = int(height * scale)

    return image, buffer


EXPORT_STATIC_TABLES: Sequence[Tuple[str, Type[ProjectLinkedMixin]]] = [
    ("Project Details", ProjectDetailsTable),
    ("Revision History", RevisionHistoryTable),
    ("Table Of Contents", TOCEntryTable),
    ("Definitions & Acronyms", DefinitionAcronymTable),
    ("Assumptions", AssumptionTable),
    ("Constraints", ConstraintTable),
    ("Dependencies", DependencyTable),
    ("Stakeholders", StakeholderTable),
    ("Deliverables", DeliverableTable),
    ("Milestone Columns", MilestoneColumnTable),
    ("SAM Deliverables", SamDeliverableTable),
    ("SAM Milestone Columns", SamMilestoneColumnTable),
]


TABLES_TO_PURGE: List[Type[ProjectLinkedMixin]] = [
    RevisionHistoryTable,
    TOCEntryTable,
    DefinitionAcronymTable,
    SingleEntryFieldTable,
    ProjectDetailsTable,
    AssumptionTable,
    ConstraintTable,
    DependencyTable,
    StakeholderTable,
    DeliverableTable,
    MilestoneColumnTable,
    SamDeliverableTable,
    SamMilestoneColumnTable,
]

TABLES_TO_PURGE.extend(
    meta.model for meta in SECTION_TABLE_REGISTRY.values()
)


def resolve_section_table(section: str, table_name: str) -> SectionTableMeta:
    try:
        return SECTION_TABLE_REGISTRY[(section, table_name)]
    except KeyError as exc:  # pragma: no cover - defensive programming
        raise HTTPException(status_code=404, detail="Table not found") from exc


def serialize_section_row(
    section: str, table_name: str, meta: SectionTableMeta, row: Any
) -> GenericTableRow:
    data = {column: getattr(row, column) for column in meta.columns}
    return GenericTableRow(
        id=row.id,
        project_id=row.project_id,
        section=section,
        table_name=table_name,
        data=data,
    )

# ==================== SECURITY ====================

SECRET_KEY = os.environ.get("SECRET_KEY", "change-this-secret")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440
SESSION_IDLE_TIMEOUT_MINUTES = int(os.environ.get("SESSION_IDLE_TIMEOUT_MINUTES", "30"))
MAX_CONCURRENT_SESSIONS = max(100, int(os.environ.get("MAX_CONCURRENT_SESSIONS", "1000")))

PORTAL_SESSION_TIMEOUT_SECONDS = float(
    os.environ.get("PORTAL_SESSION_TIMEOUT_SECONDS", "5")
)
PORTAL_BASE_URL = os.environ.get("PORTAL_BASE_URL")
PORTAL_VALIDATE_URL = os.environ.get("PORTAL_SESSION_VALIDATE_URL")
PORTAL_AUTH_ENABLED = os.environ.get("ENABLE_PORTAL_AUTH_BRIDGE", "true").lower() != "false"


security = HTTPBearer()


@dataclass
class SessionInfo:
    user_id: str
    token: str
    expires_at: datetime
    last_seen: datetime


@dataclass
class PortalUserInfo:
    email: str
    username: str
    display_name: str
    role: str


@dataclass
class PortalSessionInfo:
    token: str
    user: PortalUserInfo
    issued_at: datetime
    expires_at: datetime


@dataclass
class PortalSessionCacheEntry:
    user_id: str
    expires_at: datetime


_session_registry: Dict[str, SessionInfo] = {}
_revoked_tokens: Dict[str, datetime] = {}
_portal_session_cache: Dict[str, PortalSessionCacheEntry] = {}
_session_lock = asyncio.Lock()


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _idle_deadline(info: SessionInfo) -> datetime:
    return info.last_seen + timedelta(minutes=SESSION_IDLE_TIMEOUT_MINUTES)


def _remember_revoked_locked(token: str, expires_at: Optional[datetime], now: datetime) -> None:
    if expires_at is None:
        expires_at = now + timedelta(minutes=SESSION_IDLE_TIMEOUT_MINUTES)
    if expires_at <= now:
        return
    _revoked_tokens[token] = expires_at


def _prune_revoked_locked(now: datetime) -> None:
    expired = [token for token, expires_at in _revoked_tokens.items() if now >= expires_at]
    for token in expired:
        _revoked_tokens.pop(token, None)


async def _prune_sessions_locked(now: Optional[datetime] = None) -> None:
    if now is None:
        now = _utcnow()

    expired_tokens = [
        (token, info)
        for token, info in _session_registry.items()
        if now >= info.expires_at or now >= _idle_deadline(info)
    ]
    for token, info in expired_tokens:
        _session_registry.pop(token, None)
        _remember_revoked_locked(token, info.expires_at, now)


def _store_session_locked(info: SessionInfo, now: datetime) -> None:
    """Insert or update a session entry while respecting global caps."""

    _session_registry[info.token] = info
    _revoked_tokens.pop(info.token, None)

    if len(_session_registry) > MAX_CONCURRENT_SESSIONS:
        stalest_info = min(
            _session_registry.values(),
            key=lambda candidate: candidate.last_seen,
        )

        _session_registry.pop(stalest_info.token, None)
        _remember_revoked_locked(stalest_info.token, stalest_info.expires_at, now)


def _resolve_validate_url() -> Optional[str]:
    if PORTAL_VALIDATE_URL:
        resolved = PORTAL_VALIDATE_URL.rstrip("/")
        return f"{resolved}/"
    if PORTAL_BASE_URL:
        resolved = f"{PORTAL_BASE_URL.rstrip('/')}/api/auth/session/validate"
        return f"{resolved.rstrip('/')}/"
    return "http://localhost:9000/api/auth/session/validate/"


def _normalise_username(portal_user: PortalUserInfo) -> str:
    candidate = portal_user.username or portal_user.display_name or portal_user.email
    if not candidate:
        return "user"
    return candidate.split("@")[0]


async def _remember_portal_session(
    token: str, user_id: str, expires_at: datetime
) -> None:
    async with _session_lock:
        _portal_session_cache[token] = PortalSessionCacheEntry(
            user_id=user_id, expires_at=expires_at
        )


async def _cached_portal_user(
    token: str,
) -> Optional[Tuple[str, datetime]]:
    async with _session_lock:
        entry = _portal_session_cache.get(token)
        if entry is None:
            return None
        now = _utcnow()
        if now >= entry.expires_at:
            _portal_session_cache.pop(token, None)
            return None
        return entry.user_id, entry.expires_at


async def _ensure_portal_user(
    session: AsyncSession, portal_user: PortalUserInfo
) -> UserTable:
    result = await session.execute(
        select(UserTable).where(UserTable.email == portal_user.email)
    )
    user = result.scalar_one_or_none()

    username = _normalise_username(portal_user)
    role = portal_user.role or "user"

    if user is None:
        password_seed = secrets.token_urlsafe(24)
        user = UserTable(
            email=portal_user.email,
            username=username,
            role=role,
            password_hash=hash_password(password_seed),
        )
        session.add(user)
        await session.commit()
        await session.refresh(user)
        return user

    updated = False
    if username and user.username != username:
        user.username = username
        updated = True
    if role and user.role != role:
        user.role = role
        updated = True

    if updated:
        await session.commit()
        await session.refresh(user)

    return user


async def _fetch_portal_session(token: str) -> Optional[PortalSessionInfo]:
    if not PORTAL_AUTH_ENABLED:
        return None

    validate_url = _resolve_validate_url()
    if not validate_url:
        return None

    try:
        async with httpx.AsyncClient(
            timeout=PORTAL_SESSION_TIMEOUT_SECONDS,
            follow_redirects=False,
        ) as client:
            response = await client.get(
                validate_url,
                params={"token": token},
                headers={"Accept": "application/json"},
            )
    except httpx.HTTPError:
        return None

    if response.status_code != status.HTTP_200_OK:
        return None

    try:
        payload = response.json()
    except ValueError:
        return None

    user_payload = payload.get("user") or {}
    email = user_payload.get("email")
    if not email:
        return None
    email = str(email).strip().lower()

    try:
        issued_at_raw = int(payload.get("issued_at"))
        expires_at_raw = int(payload.get("expires_at"))
    except (TypeError, ValueError):
        return None

    issued_at = datetime.fromtimestamp(issued_at_raw, tz=timezone.utc)
    expires_at = datetime.fromtimestamp(expires_at_raw, tz=timezone.utc)

    portal_user = PortalUserInfo(
        email=email,
        username=user_payload.get("username") or "",
        display_name=user_payload.get("display_name") or "",
        role=user_payload.get("role") or "user",
    )

    return PortalSessionInfo(
        token=token,
        user=portal_user,
        issued_at=issued_at,
        expires_at=expires_at,
    )


async def authenticate_with_portal_token(
    token: str, session: AsyncSession
) -> Optional[UserProfile]:
    if not PORTAL_AUTH_ENABLED:
        return None

    cached = await _cached_portal_user(token)
    if cached:
        cached_user_id, cached_expiry = cached
        try:
            await validate_and_touch_session(
                token, cached_user_id, expires_at=cached_expiry
            )
        except HTTPException:
            async with _session_lock:
                _portal_session_cache.pop(token, None)
        else:
            user = await fetch_user_by_id(session, cached_user_id)
            if user is not None:
                return UserProfile.model_validate(user)

    portal_session = await _fetch_portal_session(token)
    if portal_session is None:
        return None

    user = await _ensure_portal_user(session, portal_session.user)

    await validate_and_touch_session(
        token, user.id, expires_at=portal_session.expires_at
    )
    await _remember_portal_session(token, user.id, portal_session.expires_at)

    return UserProfile.model_validate(user)


async def register_session(user_id: str, token: str, expires_at: datetime) -> None:
    async with _session_lock:
        now = _utcnow()
        await _prune_sessions_locked(now)
        _prune_revoked_locked(now)

        info = SessionInfo(
            user_id=user_id,
            token=token,
            expires_at=expires_at,
            last_seen=now,
        )
        _store_session_locked(info, now)


async def validate_and_touch_session(
    token: str, user_id: str, *, expires_at: Optional[datetime] = None
) -> None:
    async with _session_lock:
        now = _utcnow()
        await _prune_sessions_locked(now)
        _prune_revoked_locked(now)

        revoked_expires_at = _revoked_tokens.get(token)
        if revoked_expires_at is not None and revoked_expires_at > now:
            raise HTTPException(status_code=401, detail="Session is no longer active")
        elif revoked_expires_at is not None:
            _revoked_tokens.pop(token, None)

        info = _session_registry.get(token)
        if info is None:
            if expires_at is None:
                raise HTTPException(status_code=401, detail="Session is no longer active")

            info = SessionInfo(
                user_id=user_id,
                token=token,
                expires_at=expires_at,
                last_seen=now,
            )
            _store_session_locked(info, now)
        elif info.user_id != user_id:
            raise HTTPException(status_code=401, detail="Session is no longer active")

        if now >= info.expires_at or now >= _idle_deadline(info):
            _session_registry.pop(token, None)
            _remember_revoked_locked(token, info.expires_at, now)
            raise HTTPException(status_code=401, detail="Session has expired")

        info.last_seen = now


async def revoke_user_sessions(user_id: str) -> None:
    async with _session_lock:
        now = _utcnow()
        _prune_revoked_locked(now)
        tokens_to_remove = [
            (token, info)
            for token, info in _session_registry.items()
            if info.user_id == user_id
        ]
        for token, info in tokens_to_remove:
            _session_registry.pop(token, None)
            _portal_session_cache.pop(token, None)
            _remember_revoked_locked(token, info.expires_at, now)


async def revoke_session_by_token(token: str) -> None:
    async with _session_lock:
        now = _utcnow()
        _prune_revoked_locked(now)
        info = _session_registry.pop(token, None)
        expires_at = info.expires_at if info else None
        _portal_session_cache.pop(token, None)

        if expires_at is None:
            try:
                payload = PyJWT.decode(
                    token,
                    SECRET_KEY,
                    algorithms=[ALGORITHM],
                    options={"verify_exp": False},
                )
            except JWTError:
                payload = {}

            exp_claim = payload.get("exp")
            if isinstance(exp_claim, (int, float)):
                expires_at = datetime.fromtimestamp(exp_claim, tz=timezone.utc)

        if expires_at is not None and expires_at <= now:
            return

        _remember_revoked_locked(token, expires_at, now)

# ==================== Pydantic Schemas ====================


class UserProfile(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    username: str
    role: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class UserCreate(BaseModel):
    email: EmailStr
    username: str
    password: str
    role: str


class UserRoleUpdate(BaseModel):
    role: str


class UserInDB(UserProfile):
    password_hash: str


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserProfile


class PlanCycle(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class PlanCycleCreate(BaseModel):
    name: str
    description: Optional[str] = None


class PlanCycleUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None


class Project(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    plan_cycle_id: str
    name: str
    description: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ProjectCreate(BaseModel):
    name: str
    description: Optional[str] = None
    plan_cycle_id: str


class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None


class ProjectAccess(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    plan_cycle_id: str
    project_id: str
    visible: bool = True


class ProjectAccessUpdate(BaseModel):
    visible: bool


class RevisionHistory(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    revision_no: str
    date: str
    change_description: str
    authors: str = ""
    reviewed_by: str
    approved_by: str

    @field_validator("authors", mode="before")
    @classmethod
    def _coerce_authors(cls, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value
        return str(value)


class RevisionHistoryCreate(BaseModel):
    revision_no: str
    date: str
    change_description: str
    authors: str
    reviewed_by: str
    approved_by: str


class TOCEntry(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    sheet_name: str
    sections_in_sheet: str


class TOCEntryCreate(BaseModel):
    sheet_name: str
    sections_in_sheet: str


class DefinitionAcronym(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    term: str
    definition: str


class DefinitionAcronymCreate(BaseModel):
    term: str
    definition: str


class SingleEntryField(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    field_name: str
    content: str
    image_data: Optional[str] = None


class SingleEntryFieldCreate(BaseModel):
    field_name: str
    content: str
    image_data: Optional[str] = None


class ProjectDetails(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    project_model: str
    project_type: str
    software_type: str
    standard_to_be_followed: str
    customer: str
    programming_language: str
    project_duration: str
    team_size: str


class ProjectDetailsCreate(BaseModel):
    project_model: str
    project_type: str
    software_type: str
    standard_to_be_followed: str
    customer: str
    programming_language: str
    project_duration: str
    team_size: str


class Assumption(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    sl_no: str
    brief_description: str
    impact_on_project_objectives: str
    remarks: Optional[str] = None


class AssumptionCreate(BaseModel):
    sl_no: str
    brief_description: str
    impact_on_project_objectives: str
    remarks: Optional[str] = None


class Constraint(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    constraint_no: str
    brief_description: str
    impact_on_project_objectives: str
    remarks: Optional[str] = None


class ConstraintCreate(BaseModel):
    constraint_no: str
    brief_description: str
    impact_on_project_objectives: str
    remarks: Optional[str] = None


class Dependency(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    sl_no: str
    brief_description: str
    impact_on_project_objectives: str
    remarks: Optional[str] = None


class DependencyCreate(BaseModel):
    sl_no: str
    brief_description: str
    impact_on_project_objectives: str
    remarks: Optional[str] = None


class Stakeholder(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    sl_no: str
    name: str
    stakeholder_type: str
    role: str
    authority_responsibility: str
    contact_details: str


class StakeholderCreate(BaseModel):
    sl_no: str
    name: str
    stakeholder_type: str
    role: str
    authority_responsibility: str
    contact_details: str


class Deliverable(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    sl_no: str
    work_product: str
    owner_of_deliverable: str
    approving_authority: str
    release_to_customer: str
    milestones: Dict[str, str] = Field(default_factory=dict)


class DeliverableCreate(BaseModel):
    sl_no: str
    work_product: str
    owner_of_deliverable: str
    approving_authority: str
    release_to_customer: str
    milestones: Dict[str, str] = Field(default_factory=dict)


class MilestoneColumn(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    column_name: str
    order: int


class MilestoneColumnCreate(BaseModel):
    column_name: str


class GenericTableRow(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    section: str
    table_name: str
    data: Dict[str, Any]


class GenericTableRowCreate(BaseModel):
    data: Dict[str, Any]


class SamDeliverable(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    sl_no: str
    work_product: str
    owner_of_deliverable: str
    approving_authority: str
    release_to_tsbj: str
    milestones: Dict[str, str] = Field(default_factory=dict)


class SamDeliverableCreate(BaseModel):
    sl_no: str
    work_product: str
    owner_of_deliverable: str
    approving_authority: str
    release_to_tsbj: str
    milestones: Dict[str, str] = Field(default_factory=dict)


class SamMilestoneColumn(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    column_name: str
    order: int


class SamMilestoneColumnCreate(BaseModel):
    column_name: str


SchemaType = TypeVar("SchemaType", bound=BaseModel)
TableType = TypeVar("TableType", bound=ProjectLinkedMixin)

# ==================== UTILITIES ====================





def create_access_token(data: Dict[str, Any], expires_delta: Optional[timedelta] = None) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + (
        expires_delta if expires_delta else timedelta(minutes=15)
    )
    to_encode.update({"exp": expire})
    return PyJWT.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


async def get_session() -> AsyncGenerator[AsyncSession, None]:
    async with async_session() as session:
        yield session


async def fetch_user_by_id(session: AsyncSession, user_id: str) -> Optional[UserTable]:
    result = await session.execute(select(UserTable).where(UserTable.id == user_id))
    return result.scalar_one_or_none()


async def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    session: AsyncSession = Depends(get_session),
) -> UserProfile:
    raw_token = credentials.credentials
    try:
        payload = PyJWT.decode(raw_token, SECRET_KEY, algorithms=[ALGORITHM])
    except ExpiredSignatureError:
        await revoke_session_by_token(raw_token)
        raise HTTPException(status_code=401, detail="Token has expired")
    except JWTError:
        portal_user = await authenticate_with_portal_token(raw_token, session)
        if portal_user is not None:
            return portal_user
        await revoke_session_by_token(raw_token)
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")

    user_id: Optional[str] = payload.get("sub")
    if user_id is None:
        await revoke_session_by_token(raw_token)
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")

    exp_claim = payload.get("exp")
    expires_at = (
        datetime.fromtimestamp(exp_claim, tz=timezone.utc)
        if isinstance(exp_claim, (int, float))
        else None
    )

    await validate_and_touch_session(raw_token, user_id, expires_at=expires_at)

    user = await fetch_user_by_id(session, user_id)
    if user is None:
        await revoke_session_by_token(raw_token)
        raise HTTPException(status_code=401, detail="User not found")

    return UserProfile.model_validate(user)


async def require_admin(current_user: UserProfile = Depends(get_current_user)) -> UserProfile:
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user


async def require_editor(current_user: UserProfile = Depends(get_current_user)) -> UserProfile:
    if current_user.role not in {"admin", "editor"}:
        raise HTTPException(status_code=403, detail="Editor access required")
    return current_user


async def migrate_existing_password_hashes() -> None:
    async with async_session() as session:
        result = await session.execute(select(UserTable))
        users = result.scalars().all()
        updated = False
        for user in users:
            if not is_password_hash(user.password_hash):
                user.password_hash = hash_password(user.password_hash)
                updated = True
        if updated:
            await session.commit()


async def init_default_users() -> None:
    async with async_session() as session:
        for email, username, role, password in (
            ("admin@plankit.com", "Admin User", "admin", "admin123"),
            ("editor@plankit.com", "Editor User", "editor", "editor123"),
            ("viewer@plankit.com", "Viewer User", "viewer", "viewer123"),
        ):
            result = await session.execute(select(UserTable).where(UserTable.email == email))
            if result.scalar_one_or_none() is None:
                user = UserTable(
                    email=email,
                    username=username,
                    role=role,
                    password_hash=hash_password(password),
                )
                session.add(user)
        await session.commit()


async def get_plan_cycle_or_404(session: AsyncSession, plan_cycle_id: str) -> PlanCycleTable:
    result = await session.execute(
        select(PlanCycleTable).where(PlanCycleTable.id == plan_cycle_id)
    )
    plan_cycle = result.scalar_one_or_none()
    if plan_cycle is None:
        raise HTTPException(status_code=404, detail="Plan cycle not found")
    return plan_cycle


async def get_project_or_404(
    session: AsyncSession,
    plan_cycle_id: str,
    project_id: str,
    current_user: Optional["UserProfile"] = None,
) -> ProjectTable:
    result = await session.execute(
        select(ProjectTable).where(
            ProjectTable.id == project_id,
            ProjectTable.plan_cycle_id == plan_cycle_id,
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    if current_user is not None and current_user.role != "admin":
        access_result = await session.execute(
            select(ProjectAccessTable.visible).where(
                ProjectAccessTable.user_id == current_user.id,
                ProjectAccessTable.plan_cycle_id == plan_cycle_id,
                ProjectAccessTable.project_id == project_id,
            )
        )
        visible_override = access_result.scalar_one_or_none()
        if visible_override is False:
            raise HTTPException(status_code=404, detail="Project not found")

    return project


async def get_item_or_404(
    session: AsyncSession,
    table: Type[TableType],
    item_id: str,
    plan_cycle_id: Optional[str] = None,
    project_id: Optional[str] = None,
) -> TableType:
    stmt = select(table).where(table.id == item_id)
    if hasattr(table, "plan_cycle_id") and plan_cycle_id is not None:
        stmt = stmt.where(table.plan_cycle_id == plan_cycle_id)
    if hasattr(table, "project_id") and project_id is not None:
        stmt = stmt.where(table.project_id == project_id)
    result = await session.execute(stmt)
    item = result.scalar_one_or_none()
    if item is None:
        raise HTTPException(status_code=404, detail="Item not found")
    return item


def to_schema(schema: Type[SchemaType], instance: Any) -> SchemaType:
    return schema.model_validate(instance)


async def purge_project_children(
    session: AsyncSession, plan_cycle_id: str, project_id: str
) -> None:
    for table in TABLES_TO_PURGE:
        filters = [table.plan_cycle_id == plan_cycle_id, table.project_id == project_id]
        await session.execute(delete(table).where(*filters))
    await session.execute(
        delete(ProjectAccessTable).where(
            ProjectAccessTable.plan_cycle_id == plan_cycle_id,
            ProjectAccessTable.project_id == project_id,
        )
    )
    await session.commit()


def _apply_schema_migrations(sync_conn) -> None:
    inspector = inspect(sync_conn)
    tables = inspector.get_table_names()
    if "tailoring_qms" in tables:
        columns = {column["name"] for column in inspector.get_columns("tailoring_qms")}
        if "gap_analysis_details" not in columns:
            sync_conn.execute(
                text("ALTER TABLE tailoring_qms ADD COLUMN gap_analysis_details TEXT")
            )

    required_not_null_columns = {
        "quantitative_objectives_measurement_and_data_management_plan": [
            "sl_no",
        ],
    }

    for table_name, columns in required_not_null_columns.items():
        if table_name not in tables:
            continue
        existing_columns = {
            column["name"] for column in inspector.get_columns(table_name)
        }
        for column_name in columns:
            if column_name in existing_columns:
                continue
            sync_conn.execute(
                text(
                    f"ALTER TABLE {table_name} ADD COLUMN {column_name} TEXT "
                    "DEFAULT '' NOT NULL"
                )
            )

    default_plan_cycle_id = "default-plan-cycle"
    if "plan_cycles" in tables:
        sync_conn.execute(
            text(
                """
                INSERT OR IGNORE INTO plan_cycles (id, name, description, created_at)
                VALUES (:id, :name, :description, :created_at)
                """
            ),
            {
                "id": default_plan_cycle_id,
                "name": "Legacy Plan Cycle",
                "description": "Automatically created during migration",
                "created_at": datetime.now(timezone.utc).isoformat(),
            },
        )

    tables_with_plan_cycle = [
        table.name
        for table in Base.metadata.sorted_tables
        if "plan_cycle_id" in table.c
    ]

    for table_name in tables_with_plan_cycle:
        if table_name not in tables:
            continue
        columns = {column["name"] for column in inspector.get_columns(table_name)}
        if "plan_cycle_id" not in columns:
            sync_conn.execute(
                text(f"ALTER TABLE {table_name} ADD COLUMN plan_cycle_id TEXT")
            )
        sync_conn.execute(
            text(
                f"UPDATE {table_name} SET plan_cycle_id = :plan_cycle_id "
                "WHERE plan_cycle_id IS NULL OR plan_cycle_id = ''"
            ),
            {"plan_cycle_id": default_plan_cycle_id},
        )

    def _get_sqlite_type_and_default(column: ColumnDefinition) -> Tuple[str, Optional[Any]]:
        type_ = column.type_
        if type_ in (String, Text):
            return "TEXT", ""
        if type_ is Integer:
            return "INTEGER", 0
        if type_ is Boolean:
            return "INTEGER", 0
        if type_ is DateTime:
            return "TEXT", ""
        if type_ is JSON:
            return "TEXT", "{}"
        return "TEXT", ""

    section_tables = {
        (definition.table_name or definition.key): definition
        for definition in SECTION_TABLE_DEFINITIONS
    }

    for table_name, definition in section_tables.items():
        if table_name not in tables:
            continue
        columns = {column_info["name"] for column_info in inspector.get_columns(table_name)}
        for column in definition.columns:
            if column.name in columns:
                continue
            column_type, default_value = _get_sqlite_type_and_default(column)
            column_sql = f"{column.name} {column_type}"
            default_literal: Optional[str] = None
            if default_value is not None:
                if isinstance(default_value, str):
                    escaped_default = default_value.replace("'", "''")
                    default_literal = f"'{escaped_default}'"
                else:
                    default_literal = str(default_value)
            if not column.nullable:
                if default_literal is not None:
                    column_sql += f" DEFAULT {default_literal}"
                column_sql += " NOT NULL"
            sync_conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_sql}"))
            if not column.nullable and default_value is not None:
                sync_conn.execute(
                    text(
                        f"UPDATE {table_name} SET {column.name} = :default "
                        f"WHERE {column.name} IS NULL"
                    ),
                    {"default": default_value},
                )

    if "revision_history" in tables:
        columns = {column["name"] for column in inspector.get_columns("revision_history")}
        if "authors" not in columns:
            sync_conn.execute(
                text("ALTER TABLE revision_history ADD COLUMN authors TEXT")
            )
            columns.add("authors")
        if "remarks" in columns and "authors" in columns:
            sync_conn.execute(
                text(
                    "UPDATE revision_history SET authors = remarks "
                    "WHERE (authors IS NULL OR authors = '') "
                    "AND remarks IS NOT NULL AND remarks != ''"
                )
            )
        sync_conn.execute(
            text(
                "UPDATE revision_history SET authors = '' "
                "WHERE authors IS NULL"
            )
        )


# ==================== FASTAPI SETUP ====================


@app.on_event("startup")
async def on_startup() -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        await conn.run_sync(_apply_schema_migrations)
    await migrate_existing_password_hashes()
    await init_default_users()


@app.on_event("shutdown")
async def on_shutdown() -> None:
    await engine.dispose()


# ==================== AUTH ROUTES ====================


@api_router.post("/auth/login", response_model=Token)
async def login(login_data: LoginRequest, session: AsyncSession = Depends(get_session)) -> Token:
    result = await session.execute(select(UserTable).where(UserTable.email == login_data.email))
    user = result.scalar_one_or_none()
    if user is None or not verify_password(login_data.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Incorrect email or password")

    expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token({"sub": user.id}, expires)
    expires_at = datetime.now(timezone.utc) + expires
    await register_session(user.id, access_token, expires_at)
    return Token(access_token=access_token, token_type="bearer", user=to_schema(UserProfile, user))


@api_router.get("/auth/me", response_model=UserProfile)
async def get_current_user_info(current_user: UserProfile = Depends(get_current_user)) -> UserProfile:
    return current_user


# ==================== USER MANAGEMENT ====================


@api_router.post("/users", response_model=UserProfile)
async def create_user(
    user: UserCreate,
    current_user: UserProfile = Depends(require_admin),
    session: AsyncSession = Depends(get_session),
) -> UserProfile:
    existing = await session.execute(select(UserTable).where(UserTable.email == user.email))
    if existing.scalar_one_or_none() is not None:
        raise HTTPException(status_code=400, detail="Email already registered")

    user_in_db = UserTable(
        email=user.email,
        username=user.username,
        role=user.role,
        password_hash=hash_password(user.password),
    )
    session.add(user_in_db)
    await session.commit()
    await session.refresh(user_in_db)
    return to_schema(UserProfile, user_in_db)


@api_router.get("/users", response_model=List[UserProfile])
async def get_all_users(
    current_user: UserProfile = Depends(require_admin),
    session: AsyncSession = Depends(get_session),
) -> List[UserProfile]:
    result = await session.execute(select(UserTable))
    return [to_schema(UserProfile, row) for row in result.scalars().all()]


@api_router.patch("/users/{user_id}/role", response_model=UserProfile)
async def update_user_role(
    user_id: str,
    payload: UserRoleUpdate,
    current_user: UserProfile = Depends(require_admin),
    session: AsyncSession = Depends(get_session),
) -> UserProfile:
    user = await fetch_user_by_id(session, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")

    user.role = payload.role
    await session.commit()
    await session.refresh(user)
    return to_schema(UserProfile, user)


@api_router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    current_user: UserProfile = Depends(require_admin),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")

    user = await fetch_user_by_id(session, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")

    await revoke_user_sessions(user_id)
    await session.execute(
        delete(ProjectAccessTable).where(ProjectAccessTable.user_id == user_id)
    )
    await session.delete(user)
    await session.commit()
    return {"message": "User deleted successfully"}


@api_router.get(
    "/users/{user_id}/plan-cycles/{plan_cycle_id}/project-access",
    response_model=List[ProjectAccess],
)
async def list_user_project_access(
    user_id: str,
    plan_cycle_id: str,
    current_user: UserProfile = Depends(require_admin),
    session: AsyncSession = Depends(get_session),
) -> List[ProjectAccess]:
    user = await fetch_user_by_id(session, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")

    await get_plan_cycle_or_404(session, plan_cycle_id)
    result = await session.execute(
        select(ProjectAccessTable).where(
            ProjectAccessTable.user_id == user_id,
            ProjectAccessTable.plan_cycle_id == plan_cycle_id,
        )
    )
    return [to_schema(ProjectAccess, row) for row in result.scalars().all()]


@api_router.put(
    "/users/{user_id}/plan-cycles/{plan_cycle_id}/project-access/{project_id}",
    response_model=ProjectAccess,
)
async def upsert_project_access(
    user_id: str,
    plan_cycle_id: str,
    project_id: str,
    payload: ProjectAccessUpdate,
    current_user: UserProfile = Depends(require_admin),
    session: AsyncSession = Depends(get_session),
) -> ProjectAccess:
    user = await fetch_user_by_id(session, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")

    await get_project_or_404(session, plan_cycle_id, project_id)

    result = await session.execute(
        select(ProjectAccessTable).where(
            ProjectAccessTable.user_id == user_id,
            ProjectAccessTable.plan_cycle_id == plan_cycle_id,
            ProjectAccessTable.project_id == project_id,
        )
    )
    entry = result.scalar_one_or_none()

    if entry is None:
        entry = ProjectAccessTable(
            user_id=user_id,
            plan_cycle_id=plan_cycle_id,
            project_id=project_id,
            visible=payload.visible,
        )
        session.add(entry)
    else:
        entry.visible = payload.visible

    await session.commit()
    await session.refresh(entry)
    return to_schema(ProjectAccess, entry)


@api_router.delete(
    "/users/{user_id}/plan-cycles/{plan_cycle_id}/project-access"
)
async def reset_project_access(
    user_id: str,
    plan_cycle_id: str,
    current_user: UserProfile = Depends(require_admin),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    user = await fetch_user_by_id(session, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")

    await get_plan_cycle_or_404(session, plan_cycle_id)
    await session.execute(
        delete(ProjectAccessTable).where(
            ProjectAccessTable.user_id == user_id,
            ProjectAccessTable.plan_cycle_id == plan_cycle_id,
        )
    )
    await session.commit()
    return {"message": "Project access reset"}


# ==================== PLAN CYCLE ROUTES ====================


@api_router.post("/plan-cycles", response_model=PlanCycle)
async def create_plan_cycle(
    payload: PlanCycleCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> PlanCycle:
    plan_cycle = PlanCycleTable(name=payload.name, description=payload.description)
    session.add(plan_cycle)
    try:
        await session.commit()
    except IntegrityError as exc:
        await session.rollback()
        raise HTTPException(status_code=400, detail="Plan cycle already exists") from exc
    await session.refresh(plan_cycle)
    return to_schema(PlanCycle, plan_cycle)


@api_router.get("/plan-cycles", response_model=List[PlanCycle])
async def list_plan_cycles(
    _: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[PlanCycle]:
    result = await session.execute(select(PlanCycleTable).order_by(PlanCycleTable.created_at.desc()))
    return [to_schema(PlanCycle, row) for row in result.scalars().all()]


@api_router.get("/plan-cycles/{plan_cycle_id}", response_model=PlanCycle)
async def get_plan_cycle(
    plan_cycle_id: str,
    _: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> PlanCycle:
    plan_cycle = await get_plan_cycle_or_404(session, plan_cycle_id)
    return to_schema(PlanCycle, plan_cycle)


@api_router.put("/plan-cycles/{plan_cycle_id}", response_model=PlanCycle)
async def update_plan_cycle(
    plan_cycle_id: str,
    payload: PlanCycleUpdate,
    _: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> PlanCycle:
    plan_cycle = await get_plan_cycle_or_404(session, plan_cycle_id)
    updates = payload.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")

    for key, value in updates.items():
        setattr(plan_cycle, key, value)

    try:
        await session.commit()
    except IntegrityError as exc:
        await session.rollback()
        raise HTTPException(status_code=400, detail="Plan cycle already exists") from exc

    await session.refresh(plan_cycle)
    return to_schema(PlanCycle, plan_cycle)


# ==================== PROJECT ROUTES ====================


@api_router.post("/plan-cycles/{plan_cycle_id}/projects", response_model=Project)
async def create_project(
    plan_cycle_id: str,
    project: ProjectCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Project:
    await get_plan_cycle_or_404(session, plan_cycle_id)
    if project.plan_cycle_id and project.plan_cycle_id != plan_cycle_id:
        raise HTTPException(status_code=400, detail="Plan cycle mismatch")
    project_obj = ProjectTable(
        plan_cycle_id=plan_cycle_id,
        name=project.name,
        description=project.description,
        created_by=current_user.id,
    )
    session.add(project_obj)
    await session.commit()
    await session.refresh(project_obj)
    return to_schema(Project, project_obj)


@api_router.get("/plan-cycles/{plan_cycle_id}/projects", response_model=List[Project])
async def get_projects(
    plan_cycle_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[Project]:
    await get_plan_cycle_or_404(session, plan_cycle_id)
    stmt = select(ProjectTable).where(ProjectTable.plan_cycle_id == plan_cycle_id)

    if current_user.role != "admin":
        hidden_stmt = select(ProjectAccessTable.project_id).where(
            ProjectAccessTable.user_id == current_user.id,
            ProjectAccessTable.plan_cycle_id == plan_cycle_id,
            ProjectAccessTable.visible.is_(False),
        )
        hidden_result = await session.execute(hidden_stmt)
        hidden_ids = [row[0] for row in hidden_result.all()]
        if hidden_ids:
            stmt = stmt.where(ProjectTable.id.notin_(hidden_ids))

    result = await session.execute(stmt)
    return [to_schema(Project, row) for row in result.scalars().all()]


@api_router.get(
    "/plan-cycles/{plan_cycle_id}/projects/{project_id}", response_model=Project
)
async def get_project(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> Project:
    project = await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    return to_schema(Project, project)


@api_router.put("/plan-cycles/{plan_cycle_id}/projects/{project_id}", response_model=Project)
async def update_project(
    plan_cycle_id: str,
    project_id: str,
    payload: ProjectUpdate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Project:
    project = await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    updates = payload.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")

    for key, value in updates.items():
        setattr(project, key, value)

    await session.commit()
    await session.refresh(project)
    return to_schema(Project, project)


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}")
async def delete_project(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    project = await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    await session.delete(project)
    await purge_project_children(session, plan_cycle_id, project_id)
    return {"message": "Project deleted successfully"}


# ==================== SHARED CRUD HELPERS ====================


async def create_project_item(
    session: AsyncSession,
    table: Type[TableType],
    schema: Type[SchemaType],
    plan_cycle_id: str,
    project_id: str,
    payload: BaseModel,
    current_user: Optional["UserProfile"] = None,
) -> SchemaType:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    obj = table(
        plan_cycle_id=plan_cycle_id,
        project_id=project_id,
        **payload.model_dump(),
    )
    session.add(obj)
    await session.commit()
    await session.refresh(obj)
    return to_schema(schema, obj)


async def list_project_items(
    session: AsyncSession,
    table: Type[TableType],
    schema: Type[SchemaType],
    plan_cycle_id: str,
    project_id: str,
    order_by: Optional[Any] = None,
    current_user: Optional["UserProfile"] = None,
) -> List[SchemaType]:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    stmt = select(table).where(
        table.plan_cycle_id == plan_cycle_id, table.project_id == project_id
    )
    if order_by is not None:
        stmt = stmt.order_by(order_by)
    result = await session.execute(stmt)
    return [to_schema(schema, row) for row in result.scalars().all()]


async def update_project_item(
    session: AsyncSession,
    table: Type[TableType],
    schema: Type[SchemaType],
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    payload: BaseModel,
    extra_updates: Optional[Dict[str, Any]] = None,
    current_user: Optional["UserProfile"] = None,
) -> SchemaType:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    obj = await get_item_or_404(
        session,
        table,
        item_id,
        plan_cycle_id=plan_cycle_id,
        project_id=project_id,
    )
    data = payload.model_dump()
    if extra_updates:
        data.update(extra_updates)
    for key, value in data.items():
        setattr(obj, key, value)
    await session.commit()
    await session.refresh(obj)
    return to_schema(schema, obj)


async def delete_project_item(
    session: AsyncSession,
    table: Type[TableType],
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    current_user: Optional["UserProfile"] = None,
) -> Dict[str, str]:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    obj = await get_item_or_404(
        session,
        table,
        item_id,
        plan_cycle_id=plan_cycle_id,
        project_id=project_id,
    )
    await session.delete(obj)
    await session.commit()
    return {"message": "Item deleted successfully"}


# ==================== MODULE ENDPOINTS ====================


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/revision-history", response_model=RevisionHistory)
async def create_revision_history(
    plan_cycle_id: str,
    project_id: str,
    item: RevisionHistoryCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> RevisionHistory:
    return await create_project_item(
        session,
        RevisionHistoryTable,
        RevisionHistory,
        plan_cycle_id,
        project_id,
        item,
        current_user=current_user,
    )


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/revision-history", response_model=List[RevisionHistory])
async def get_revision_history(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[RevisionHistory]:
    return await list_project_items(
        session,
        RevisionHistoryTable,
        RevisionHistory,
        plan_cycle_id,
        project_id,
        current_user=current_user,
    )


@api_router.put("/plan-cycles/{plan_cycle_id}/projects/{project_id}/revision-history/{item_id}", response_model=RevisionHistory)
async def update_revision_history(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    item: RevisionHistoryCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> RevisionHistory:
    return await update_project_item(
        session,
        RevisionHistoryTable,
        RevisionHistory,
        plan_cycle_id,
        project_id,
        item_id,
        item,
        current_user=current_user,
    )


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}/revision-history/{item_id}")
async def delete_revision_history(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    return await delete_project_item(
        session,
        RevisionHistoryTable,
        plan_cycle_id,
        project_id,
        item_id,
        current_user=current_user,
    )


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/toc-entries", response_model=TOCEntry)
async def create_toc_entry(
    plan_cycle_id: str,
    project_id: str,
    item: TOCEntryCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> TOCEntry:
    return await create_project_item(
        session,
        TOCEntryTable,
        TOCEntry,
        plan_cycle_id,
        project_id,
        item,
        current_user=current_user,
    )


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/toc-entries", response_model=List[TOCEntry])
async def get_toc_entries(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[TOCEntry]:
    return await list_project_items(
        session,
        TOCEntryTable,
        TOCEntry,
        plan_cycle_id,
        project_id,
        current_user=current_user,
    )


@api_router.put("/plan-cycles/{plan_cycle_id}/projects/{project_id}/toc-entries/{item_id}", response_model=TOCEntry)
async def update_toc_entry(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    item: TOCEntryCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> TOCEntry:
    return await update_project_item(
        session,
        TOCEntryTable,
        TOCEntry,
        plan_cycle_id,
        project_id,
        item_id,
        item,
        current_user=current_user,
    )


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}/toc-entries/{item_id}")
async def delete_toc_entry(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    return await delete_project_item(
        session,
        TOCEntryTable,
        plan_cycle_id,
        project_id,
        item_id,
        current_user=current_user,
    )


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/definition-acronyms", response_model=DefinitionAcronym)
async def create_definition_acronym(
    plan_cycle_id: str,
    project_id: str,
    item: DefinitionAcronymCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> DefinitionAcronym:
    return await create_project_item(
        session,
        DefinitionAcronymTable,
        DefinitionAcronym,
        plan_cycle_id,
        project_id,
        item,
        current_user=current_user,
    )


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/definition-acronyms", response_model=List[DefinitionAcronym])
async def get_definition_acronyms(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[DefinitionAcronym]:
    return await list_project_items(
        session,
        DefinitionAcronymTable,
        DefinitionAcronym,
        plan_cycle_id,
        project_id,
        current_user=current_user,
    )


@api_router.put("/plan-cycles/{plan_cycle_id}/projects/{project_id}/definition-acronyms/{item_id}", response_model=DefinitionAcronym)
async def update_definition_acronym(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    item: DefinitionAcronymCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> DefinitionAcronym:
    return await update_project_item(
        session,
        DefinitionAcronymTable,
        DefinitionAcronym,
        plan_cycle_id,
        project_id,
        item_id,
        item,
        current_user=current_user,
    )


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}/definition-acronyms/{item_id}")
async def delete_definition_acronym(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    return await delete_project_item(
        session,
        DefinitionAcronymTable,
        plan_cycle_id,
        project_id,
        item_id,
        current_user=current_user,
    )


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/single-entry", response_model=SingleEntryField)
async def create_or_update_single_entry(
    plan_cycle_id: str,
    project_id: str,
    item: SingleEntryFieldCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> SingleEntryField:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    stmt = select(SingleEntryFieldTable).where(
        SingleEntryFieldTable.plan_cycle_id == plan_cycle_id,
        SingleEntryFieldTable.project_id == project_id,
        SingleEntryFieldTable.field_name == item.field_name,
    )
    existing = await session.execute(stmt)
    row = existing.scalar_one_or_none()
    if row is None:
        new_item = SingleEntryFieldTable(
            plan_cycle_id=plan_cycle_id,
            project_id=project_id,
            **item.model_dump(),
        )
        session.add(new_item)
        await session.commit()
        await session.refresh(new_item)
        return to_schema(SingleEntryField, new_item)

    row.content = item.content
    row.image_data = item.image_data
    await session.commit()
    await session.refresh(row)
    return to_schema(SingleEntryField, row)


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/single-entry/{field_name}", response_model=Optional[SingleEntryField])
async def get_single_entry(
    plan_cycle_id: str,
    project_id: str,
    field_name: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> Optional[SingleEntryField]:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    stmt = select(SingleEntryFieldTable).where(
        SingleEntryFieldTable.plan_cycle_id == plan_cycle_id,
        SingleEntryFieldTable.project_id == project_id,
        SingleEntryFieldTable.field_name == field_name,
    )
    result = await session.execute(stmt)
    row = result.scalar_one_or_none()
    return to_schema(SingleEntryField, row) if row else None


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/export/xlsx")
async def export_project_xlsx(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    project = await get_project_or_404(session, plan_cycle_id, project_id, current_user)

    workbook = Workbook()
    used_titles: Set[str] = set()
    retained_image_streams: List[BytesIO] = []

    summary_sheet = workbook.active
    summary_sheet.title = make_sheet_title("Project Summary", used_titles)
    summary_sheet.column_dimensions["A"].width = 20
    summary_sheet.column_dimensions["B"].width = 80

    created_at = getattr(project, "created_at", None)
    summary_rows = [
        ("Project ID", project.id),
        ("Project Name", project.name),
        ("Description", project.description or ""),
        ("Created By", project.created_by),
        (
            "Created At",
            created_at.isoformat() if isinstance(created_at, datetime) else "",
        ),
    ]

    for label, value in summary_rows:
        summary_sheet.append([label, value])

    single_entry_stmt = select(SingleEntryFieldTable).where(
        SingleEntryFieldTable.plan_cycle_id == plan_cycle_id,
        SingleEntryFieldTable.project_id == project_id,
    )
    single_entries = (await session.execute(single_entry_stmt)).scalars().all()
    if single_entries:
        single_sheet = workbook.create_sheet(
            title=make_sheet_title("Single Entries", used_titles)
        )
        single_sheet.append(["Field Name", "Content", "Image"])
        single_sheet.column_dimensions["A"].width = 32
        single_sheet.column_dimensions["B"].width = 80
        single_sheet.column_dimensions["C"].width = 50

        row_index = 2
        for entry in sorted(single_entries, key=lambda item: item.field_name):
            single_sheet.cell(row=row_index, column=1, value=entry.field_name)
            single_sheet.cell(row=row_index, column=2, value=entry.content)
            if entry.image_data:
                decoded = decode_image_for_workbook(entry.image_data)
                if decoded is not None:
                    image, buffer = decoded
                    single_sheet.add_image(image, f"C{row_index}")
                    retained_image_streams.append(buffer)
                    if getattr(image, "height", None):
                        single_sheet.row_dimensions[row_index].height = max(
                            single_sheet.row_dimensions[row_index].height or 15,
                            image.height * 0.75,
                        )
                else:
                    single_sheet.cell(
                        row=row_index,
                        column=3,
                        value="Image unavailable",
                    )
            row_index += 1

    for sheet_title, model in EXPORT_STATIC_TABLES:
        stmt = select(model).where(
            model.plan_cycle_id == plan_cycle_id, model.project_id == project_id
        )
        rows = (await session.execute(stmt)).scalars().all()
        if not rows:
            continue

        columns = [
            column.name
            for column in model.__table__.columns
            if column.name not in EXPORT_COLUMN_EXCLUDES
        ]
        if not columns:
            continue

        sheet = workbook.create_sheet(title=make_sheet_title(sheet_title, used_titles))
        sheet.append([friendly_header(column) for column in columns])

        for idx, _ in enumerate(columns, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = 24

        for row in rows:
            sheet.append([format_cell_value(getattr(row, column)) for column in columns])

    for (section, table_name), meta in SECTION_TABLE_REGISTRY.items():
        stmt = select(meta.model).where(
            meta.model.plan_cycle_id == plan_cycle_id,
            meta.model.project_id == project_id,
        )
        rows = (await session.execute(stmt)).scalars().all()
        if not rows:
            continue

        base_title = f"{section} {table_name.replace('_', ' ').title()}"
        sheet = workbook.create_sheet(title=make_sheet_title(base_title, used_titles))
        sheet.append([friendly_header(column) for column in meta.columns])

        for idx, _ in enumerate(meta.columns, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = 24

        for row in rows:
            sheet.append(
                [format_cell_value(getattr(row, column)) for column in meta.columns]
            )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    safe_name = "".join(
        char if char.isalnum() else "_" for char in (project.name or project.id)
    ).strip("_")
    filename = f"{safe_name or project.id}.xlsx"

    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
    }

    return StreamingResponse(
        output,
        media_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/project-details", response_model=ProjectDetails)
async def create_project_details(
    plan_cycle_id: str,
    project_id: str,
    item: ProjectDetailsCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> ProjectDetails:
    return await create_project_item(session, ProjectDetailsTable, ProjectDetails, plan_cycle_id, project_id, item, current_user=current_user)


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/project-details", response_model=Optional[ProjectDetails])
async def get_project_details(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> Optional[ProjectDetails]:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    stmt = select(ProjectDetailsTable).where(
        ProjectDetailsTable.plan_cycle_id == plan_cycle_id,
        ProjectDetailsTable.project_id == project_id,
    )
    result = await session.execute(stmt)
    row = result.scalar_one_or_none()
    return to_schema(ProjectDetails, row) if row else None


@api_router.put("/plan-cycles/{plan_cycle_id}/projects/{project_id}/project-details/{item_id}", response_model=ProjectDetails)
async def update_project_details(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    item: ProjectDetailsCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> ProjectDetails:
    return await update_project_item(session, ProjectDetailsTable, ProjectDetails, plan_cycle_id, project_id, item_id, item, current_user=current_user)


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/assumptions", response_model=Assumption)
async def create_assumption(
    plan_cycle_id: str,
    project_id: str,
    item: AssumptionCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Assumption:
    return await create_project_item(session, AssumptionTable, Assumption, plan_cycle_id, project_id, item, current_user=current_user)


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/assumptions", response_model=List[Assumption])
async def get_assumptions(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[Assumption]:
    return await list_project_items(session, AssumptionTable, Assumption, plan_cycle_id, project_id, current_user=current_user)


@api_router.put("/plan-cycles/{plan_cycle_id}/projects/{project_id}/assumptions/{item_id}", response_model=Assumption)
async def update_assumption(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    item: AssumptionCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Assumption:
    return await update_project_item(session, AssumptionTable, Assumption, plan_cycle_id, project_id, item_id, item, current_user=current_user)


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}/assumptions/{item_id}")
async def delete_assumption(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    return await delete_project_item(session, AssumptionTable, plan_cycle_id, project_id, item_id, current_user=current_user)


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/constraints", response_model=Constraint)
async def create_constraint(
    plan_cycle_id: str,
    project_id: str,
    item: ConstraintCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Constraint:
    return await create_project_item(session, ConstraintTable, Constraint, plan_cycle_id, project_id, item, current_user=current_user)


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/constraints", response_model=List[Constraint])
async def get_constraints(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[Constraint]:
    return await list_project_items(session, ConstraintTable, Constraint, plan_cycle_id, project_id, current_user=current_user)


@api_router.put("/plan-cycles/{plan_cycle_id}/projects/{project_id}/constraints/{item_id}", response_model=Constraint)
async def update_constraint(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    item: ConstraintCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Constraint:
    return await update_project_item(session, ConstraintTable, Constraint, plan_cycle_id, project_id, item_id, item, current_user=current_user)


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}/constraints/{item_id}")
async def delete_constraint(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    return await delete_project_item(session, ConstraintTable, plan_cycle_id, project_id, item_id, current_user=current_user)


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/dependencies", response_model=Dependency)
async def create_dependency(
    plan_cycle_id: str,
    project_id: str,
    item: DependencyCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dependency:
    return await create_project_item(session, DependencyTable, Dependency, plan_cycle_id, project_id, item, current_user=current_user)


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/dependencies", response_model=List[Dependency])
async def get_dependencies(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[Dependency]:
    return await list_project_items(session, DependencyTable, Dependency, plan_cycle_id, project_id, current_user=current_user)


@api_router.put("/plan-cycles/{plan_cycle_id}/projects/{project_id}/dependencies/{item_id}", response_model=Dependency)
async def update_dependency(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    item: DependencyCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dependency:
    return await update_project_item(session, DependencyTable, Dependency, plan_cycle_id, project_id, item_id, item, current_user=current_user)


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}/dependencies/{item_id}")
async def delete_dependency(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    return await delete_project_item(session, DependencyTable, plan_cycle_id, project_id, item_id, current_user=current_user)


# Default stakeholder roles seeded for each project when the register is empty.
DEFAULT_STAKEHOLDER_ROLES = [
    "<SMT>",
    "<Program Manager/Delivery Head>",
    "<Customer Representative>",
    "<Engineeering Manager>",
    "<Lead Engineer/Team Leader>",
    "<Technical Leader/Technical Architect>",
    "<SSE/SE/ASE/TE>",
    "<P-SQA>",
    "<Configuration Controller>",
    "<Data Owner>",
    "<Data Custodian>",
    "<Asset Owner>",
    "<Security Leader/Officer>",
    "<Security Engineer>",
    "<Risk management Team member>",
    "<Incident Response team Member>",
    "<Vulnerability Engineer>",
    "<Red Team>",
    "<QA Team>",
    "<PV Team>",
    "<HR/Admin/Finance Teams>",
    "<IT Team>",
]


async def ensure_stakeholder_defaults(
    session: AsyncSession, plan_cycle_id: str, project_id: str
) -> None:
    count_result = await session.execute(
        select(func.count(StakeholderTable.id)).where(
            StakeholderTable.plan_cycle_id == plan_cycle_id,
            StakeholderTable.project_id == project_id,
        )
    )
    current_count = count_result.scalar_one()
    if current_count:
        return

    session.add_all(
        [
            StakeholderTable(
                plan_cycle_id=plan_cycle_id,
                project_id=project_id,
                sl_no=str(index + 1),
                name="",
                stakeholder_type="",
                role=role,
                authority_responsibility="",
                contact_details="",
            )
            for index, role in enumerate(DEFAULT_STAKEHOLDER_ROLES)
        ]
    )
    await session.commit()


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/stakeholders", response_model=Stakeholder)
async def create_stakeholder(
    plan_cycle_id: str,
    project_id: str,
    item: StakeholderCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Stakeholder:
    return await create_project_item(session, StakeholderTable, Stakeholder, plan_cycle_id, project_id, item, current_user=current_user)


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/stakeholders", response_model=List[Stakeholder])
async def get_stakeholders(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[Stakeholder]:
    await ensure_stakeholder_defaults(session, plan_cycle_id, project_id)
    return await list_project_items(session, StakeholderTable, Stakeholder, plan_cycle_id, project_id, current_user=current_user)


@api_router.put("/plan-cycles/{plan_cycle_id}/projects/{project_id}/stakeholders/{item_id}", response_model=Stakeholder)
async def update_stakeholder(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    item: StakeholderCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Stakeholder:
    return await update_project_item(session, StakeholderTable, Stakeholder, plan_cycle_id, project_id, item_id, item, current_user=current_user)


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}/stakeholders/{item_id}")
async def delete_stakeholder(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    return await delete_project_item(session, StakeholderTable, plan_cycle_id, project_id, item_id, current_user=current_user)


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/milestone-columns", response_model=MilestoneColumn)
async def create_milestone_column(
    plan_cycle_id: str,
    project_id: str,
    item: MilestoneColumnCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> MilestoneColumn:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    result = await session.execute(
        select(func.max(MilestoneColumnTable.order)).where(
            MilestoneColumnTable.plan_cycle_id == plan_cycle_id,
            MilestoneColumnTable.project_id == project_id,
        )
    )
    current_max = result.scalar_one_or_none() or 0
    column = MilestoneColumnTable(
        plan_cycle_id=plan_cycle_id,
        project_id=project_id,
        column_name=item.column_name,
        order=current_max + 1,
    )
    session.add(column)
    await session.commit()
    await session.refresh(column)
    return to_schema(MilestoneColumn, column)


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/milestone-columns", response_model=List[MilestoneColumn])
async def get_milestone_columns(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[MilestoneColumn]:
    return await list_project_items(
        session,
        MilestoneColumnTable,
        MilestoneColumn,
        plan_cycle_id,
        project_id,
        order_by=MilestoneColumnTable.order.asc(),
        current_user=current_user,
    )


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}/milestone-columns/{column_id}")
async def delete_milestone_column(
    plan_cycle_id: str,
    project_id: str,
    column_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    return await delete_project_item(session, MilestoneColumnTable, plan_cycle_id, project_id, column_id, current_user=current_user)


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/deliverables", response_model=Deliverable)
async def create_deliverable(
    plan_cycle_id: str,
    project_id: str,
    item: DeliverableCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Deliverable:
    return await create_project_item(session, DeliverableTable, Deliverable, plan_cycle_id, project_id, item, current_user=current_user)


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/deliverables", response_model=List[Deliverable])
async def get_deliverables(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[Deliverable]:
    return await list_project_items(session, DeliverableTable, Deliverable, plan_cycle_id, project_id, current_user=current_user)


@api_router.put("/plan-cycles/{plan_cycle_id}/projects/{project_id}/deliverables/{item_id}", response_model=Deliverable)
async def update_deliverable(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    item: DeliverableCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Deliverable:
    return await update_project_item(session, DeliverableTable, Deliverable, plan_cycle_id, project_id, item_id, item, current_user=current_user)


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}/deliverables/{item_id}")
async def delete_deliverable(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    return await delete_project_item(session, DeliverableTable, plan_cycle_id, project_id, item_id, current_user=current_user)


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/sam-milestone-columns", response_model=SamMilestoneColumn)
async def create_sam_milestone_column(
    plan_cycle_id: str,
    project_id: str,
    item: SamMilestoneColumnCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> SamMilestoneColumn:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    result = await session.execute(
        select(func.max(SamMilestoneColumnTable.order)).where(
            SamMilestoneColumnTable.plan_cycle_id == plan_cycle_id,
            SamMilestoneColumnTable.project_id == project_id,
        )
    )
    current_max = result.scalar_one_or_none() or 0
    column = SamMilestoneColumnTable(
        plan_cycle_id=plan_cycle_id,
        project_id=project_id,
        column_name=item.column_name,
        order=current_max + 1,
    )
    session.add(column)
    await session.commit()
    await session.refresh(column)
    return to_schema(SamMilestoneColumn, column)


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/sam-milestone-columns", response_model=List[SamMilestoneColumn])
async def get_sam_milestone_columns(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[SamMilestoneColumn]:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    stmt = (
        select(SamMilestoneColumnTable)
        .where(
            SamMilestoneColumnTable.plan_cycle_id == plan_cycle_id,
            SamMilestoneColumnTable.project_id == project_id,
        )
        .order_by(SamMilestoneColumnTable.order.asc())
    )
    result = await session.execute(stmt)
    return [to_schema(SamMilestoneColumn, row) for row in result.scalars().all()]


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}/sam-milestone-columns/{column_id}")
async def delete_sam_milestone_column(
    plan_cycle_id: str,
    project_id: str,
    column_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    return await delete_project_item(
        session,
        SamMilestoneColumnTable,
        plan_cycle_id,
        project_id,
        column_id,
        current_user=current_user,
    )


@api_router.post("/plan-cycles/{plan_cycle_id}/projects/{project_id}/sam-deliverables", response_model=SamDeliverable)
async def create_sam_deliverable(
    plan_cycle_id: str,
    project_id: str,
    item: SamDeliverableCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> SamDeliverable:
    return await create_project_item(
        session,
        SamDeliverableTable,
        SamDeliverable,
        plan_cycle_id,
        project_id,
        item,
        current_user=current_user,
    )


@api_router.get("/plan-cycles/{plan_cycle_id}/projects/{project_id}/sam-deliverables", response_model=List[SamDeliverable])
async def get_sam_deliverables(
    plan_cycle_id: str,
    project_id: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[SamDeliverable]:
    return await list_project_items(
        session,
        SamDeliverableTable,
        SamDeliverable,
        plan_cycle_id,
        project_id,
        current_user=current_user,
    )


@api_router.put(
    "/plan-cycles/{plan_cycle_id}/projects/{project_id}/sam-deliverables/{item_id}", response_model=SamDeliverable
)
async def update_sam_deliverable(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    item: SamDeliverableCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> SamDeliverable:
    return await update_project_item(
        session,
        SamDeliverableTable,
        SamDeliverable,
        plan_cycle_id,
        project_id,
        item_id,
        item,
        current_user=current_user,
    )


@api_router.delete("/plan-cycles/{plan_cycle_id}/projects/{project_id}/sam-deliverables/{item_id}")
async def delete_sam_deliverable(
    plan_cycle_id: str,
    project_id: str,
    item_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    return await delete_project_item(
        session,
        SamDeliverableTable,
        plan_cycle_id,
        project_id,
        item_id,
        current_user=current_user,
    )


@api_router.post(
    "/plan-cycles/{plan_cycle_id}/projects/{project_id}/sections/{section}/tables/{table_name}",
    response_model=GenericTableRow,
)
async def create_generic_table_row(
    plan_cycle_id: str,
    project_id: str,
    section: str,
    table_name: str,
    item: GenericTableRowCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> GenericTableRow:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    meta = resolve_section_table(section, table_name)
    row = meta.model(
        plan_cycle_id=plan_cycle_id,
        project_id=project_id,
        **{column: item.data.get(column) for column in meta.columns},
    )
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return serialize_section_row(section, table_name, meta, row)


@api_router.get(
    "/plan-cycles/{plan_cycle_id}/projects/{project_id}/sections/{section}/tables/{table_name}",
    response_model=List[GenericTableRow],
)
async def get_generic_table_rows(
    plan_cycle_id: str,
    project_id: str,
    section: str,
    table_name: str,
    current_user: UserProfile = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> List[GenericTableRow]:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    meta = resolve_section_table(section, table_name)
    stmt = select(meta.model).where(
        meta.model.plan_cycle_id == plan_cycle_id,
        meta.model.project_id == project_id,
    )
    result = await session.execute(stmt)
    return [
        serialize_section_row(section, table_name, meta, row)
        for row in result.scalars().all()
    ]


@api_router.put(
    "/plan-cycles/{plan_cycle_id}/projects/{project_id}/sections/{section}/tables/{table_name}/{item_id}",
    response_model=GenericTableRow,
)
async def update_generic_table_row(
    plan_cycle_id: str,
    project_id: str,
    section: str,
    table_name: str,
    item_id: str,
    item: GenericTableRowCreate,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> GenericTableRow:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    meta = resolve_section_table(section, table_name)
    row = await get_item_or_404(
        session,
        meta.model,
        item_id,
        plan_cycle_id=plan_cycle_id,
        project_id=project_id,
    )
    for column in meta.columns:
        setattr(row, column, item.data.get(column))
    await session.commit()
    await session.refresh(row)
    return serialize_section_row(section, table_name, meta, row)


@api_router.delete(
    "/plan-cycles/{plan_cycle_id}/projects/{project_id}/sections/{section}/tables/{table_name}/{item_id}"
)
async def delete_generic_table_row(
    plan_cycle_id: str,
    project_id: str,
    section: str,
    table_name: str,
    item_id: str,
    current_user: UserProfile = Depends(require_editor),
    session: AsyncSession = Depends(get_session),
) -> Dict[str, str]:
    await get_project_or_404(session, plan_cycle_id, project_id, current_user)
    meta = resolve_section_table(section, table_name)
    row = await get_item_or_404(
        session,
        meta.model,
        item_id,
        plan_cycle_id=plan_cycle_id,
        project_id=project_id,
    )
    await session.delete(row)
    await session.commit()
    return {"message": "Item deleted successfully"}


app.include_router(api_router)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

